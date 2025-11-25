# utils/export.py
import io
import pandas as pd
import streamlit as st 
import plotly.io as pio
import zipfile 
import openpyxl 

# Tenta importar as bibliotecas de imagem (apenas para verificação, não usado para HTML)
try:
    from openpyxl.drawing.image import Image as OpenpyxlImage # type: ignore
    from PIL import Image as PillowImage # type: ignore
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False


def create_zip_package(tables_to_export: dict, active_filters: str = "") -> bytes:
    """
    Cria um arquivo ZIP in-memory contendo o arquivo XLSX (com filtros no cabeçalho) e os gráficos HTML.
    
    Args:
        tables_to_export: Dicionário de DFs e Figuras.
        active_filters: String contendo os filtros ativos para injeção nas células A1/A2.
    
    Returns:
        Bytes do arquivo ZIP.
    """
    
    zip_buffer = io.BytesIO()
    has_real_data = False 

    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        
        # --- A. Geração do Arquivo Excel (.xlsx) ---
        excel_buffer = io.BytesIO()
        engine = 'openpyxl'
        
        try:
            import openpyxl
        except ImportError:
            st.error("A biblioteca 'openpyxl' é necessária para criar o Excel.")
            return b""

        with pd.ExcelWriter(excel_buffer, engine=engine) as writer:
            
            for sheet_name, data in tables_to_export.items():
                if data.get('df') is not None and not data['df'].empty:
                    safe_name = sheet_name.replace(":", "").replace("/", "")[:31]
                    df = data['df']
                    
                    # 1. Exporta a tabela para começar na linha 3 (startrow=2 é base 0)
                    # Isso deixa A1 e A2 livres.
                    df.to_excel(writer, sheet_name=safe_name, index=False, startrow=2) 
                    has_real_data = True
                    
                    # 2. Injeta os filtros nas células A1 e A2
                    # Acessa a worksheet do openpyxl criada pelo pandas
                    ws = writer.sheets[safe_name]
                    
                    ws['A1'] = "FILTROS ATIVOS NO MOMENTO DA EXPORTAÇÃO:"
                    # Estiliza A1 com negrito (opcional, mas bom para destaque)
                    ws['A1'].font = openpyxl.styles.Font(bold=True)
                    
                    ws['A2'] = active_filters
                    ws['A2'].alignment = openpyxl.styles.Alignment(wrap_text=False)
            
            # Garante que o ExcelWriter não falhe se não houver abas de dados
            if not has_real_data:
                info_df = pd.DataFrame({'Info': ['Este arquivo de dados não possui tabelas, pois apenas gráficos foram selecionados para exportação.']})
                info_df.to_excel(writer, sheet_name='Info_Vazio', index=False)
        
        # Define o nome do arquivo Excel no ZIP
        excel_filename = 'Dados_Tabelas.xlsx'
        
        # Só grava o Excel no ZIP se tiver dados reais. 
        # Se for só gráfico, a gente ignora o Excel gerado (que só teria a aba Info_Vazio).
        # Porém, a lógica de "deletar" o buffer é complexa. 
        # Vamos gravar o buffer SOMENTE se has_real_data for True.
        if has_real_data:
            zf.writestr(excel_filename, excel_buffer.getvalue())

        # --- B. Geração e Adição dos Gráficos HTML ao ZIP ---
        for sheet_name, data in tables_to_export.items():
            fig = data.get('fig')
            
            if fig is not None:
                try:
                    safe_name = sheet_name.replace(":", "").replace("/", "")[:31]
                    # Gera HTML completo
                    html_content = pio.to_html(fig, full_html=True, include_plotlyjs='cdn')
                    
                    file_name = f"{safe_name}_Grafico.html"
                    zf.writestr(file_name, html_content)
                
                except Exception as e:
                    st.error(f"Falha ao gerar o HTML para '{sheet_name}'. Gráfico não incluído no pacote. Erro: {e}")
        
    zip_buffer.seek(0)
    return zip_buffer.getvalue()